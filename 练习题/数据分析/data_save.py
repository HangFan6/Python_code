# -*- coding:utf-8 -*-
"""
作者：HET
日期：2023年05月06日
"""
import pandas as pd
import pandas_datareader as pdr


data=pdr.get_data_fred('GS10')
data2=pdr.get_data_fred('GS5')
data['GS5']=data2  # 值和index的精确匹配
data['mean']=(data['GS10']+data['GS5'])/2
data=round(data,2)
# print(data)

# **********CSV****************
data.to_csv('GS.csv')  # csv存储
filename='GS.csv'
gs=pd.read_csv(filename)  # csv读取,会自动创建序号
# print(gs)
gs=pd.read_csv(filename,index_col='DATE')  # 指定索引,不创建索引序号
# print(gs)

# 存储前对数据索引进行命名：date
# print(data.index.name)  # 查看索引名称 DATE
data.index.name='date'  # 修改索引名称
# print(data)
data.to_csv('GS.csv')  # csv存储
filename='GS.csv'
gs=pd.read_csv(filename)  # csv读取
# print(gs)
# # =====数据追加===========
data3=data.tail()
data3.to_csv(filename,mode='a',header=False)  # 追加数据，不追加标题


# **********EXCEL***************
data.to_excel('GS.xlsx',sheet_name='GS')  # excel存储，没有mode参数，不能进行追加
# ======一次写入多个sheet========
with pd.ExcelWriter('writer.xlsx') as writer:
    data.to_excel(writer,sheet_name='a')
    data.to_excel(writer,sheet_name='b')
    data.to_excel(writer,sheet_name='c')
# ======追加新sheet========
with pd.ExcelWriter('writer.xlsx',mode='a',engine='openpyxl') as writer:
    data3.to_excel(writer,sheet_name='d')

# ======每个sheet中追加数据==========
# 利用openpyxl库：读取sheet -> 进行追加数据 -> 存入表格
import openpyxl
wb = openpyxl.load_workbook('writer.xlsx')  # 打开excel
wb.save('writer[copy].xlsx')  # 保存文件 重命名拷贝excel文件

sheet1 = wb['d']  # 选择工作簿
sheet2 = wb.copy_worksheet(sheet1)  # 拷贝工作簿
sheet2.title = "sheet2"  # 拷贝工作簿命名
wb.save('writer.xlsx')  # 保存文件

#    ******  数据追加方法1 （列表追加）    *****
# 适用于不含特殊索引的数据追加
rows = sheet1.max_row # 读取最后一行 行号
prev_date_str = sheet1.cell(row=rows,column=1).value # 取出最后一行，第一列 日期的字符串
print(data3.values)  # 列表数据,数据类型为numpy.ndarray
lst2=data3.values.tolist()   # 用ndarray对象的方法tolist()实现 将numpy.ndarray转换为list
print(lst2)
for i in range(len(data3.values)):
    sheet1.append(lst2[i])  # 数据追加（从第一列开始追加）

# #    ******  数据追加方法2  （单元格追加）   *****
# row = sheet1.max_row   # 读取最后一行 行号
# print(row)
# for i in range(len(lst2)):
#     row += 1
#     col = 1  # 从第2列开始追加，第1列为索引号
#     for j in range(3):
#         col += 1
#         sheet1.cell(row, col, lst2[i][j])  # 数据追加
#         # 索引号为日期格式，可自行在工作簿中进行填充补充
# wb.save('writer.xlsx')  # 保存文件



